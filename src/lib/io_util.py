import openpyxl
import xlrd
import datetime
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from src.lib.log import HGJLogger

logger = HGJLogger()


def get_data_from_excel(excel_file: str, min_cols=None, min_rows=None, max_cols=None, max_rows=None,
                        sheet_index=0) -> list:
    workbook = xlrd.open_workbook(excel_file)
    sheet = workbook.sheets()[sheet_index]
    data_record = []
    upper_cols = max_cols if max_cols is not None else sheet.ncols
    lower_cols = min_cols if min_cols is not None else 0
    upper_rows = max_rows if max_rows is not None else sheet.nrows
    lower_rows = min_rows if min_rows is not None else 0
    # 双层遍历读取数据
    for row_num in range(lower_rows, upper_rows):
        row_data = list()
        for col_num in range(lower_cols, upper_cols):
            # if sheet.cell(row_num, col_num).ctype in [2, 3]:  # 检测到该格数据为日期转换成的数字
            #     tupled_date = xlrd.xldate_as_tuple(sheet.cell(row_num, col_num).value, 0)
            #     datetime_date = datetime.datetime(*tupled_date)
            #     row_data.append(datetime_date)
            # else:
                row_data.append(str(sheet.cell(row_num, col_num).value).strip())
        data_record.append(row_data)  # 结束本行数据读取，本行数据存入到数据
    return data_record


def save_data_to_excel(excel_file: str, sheet_name: str, data_seqence, title_row=None, sheet_index: int=0) -> None:
    """将文件写入excel文件
    注意：如果目标文件已存在会覆盖原来的文件。

    Parameters
    ----------
    excel_file:str: 存储的目标文件
    sheet_name:str: 存储的excel页名字
    data_seqence: 需要存储的数据，二维列表
    title_row=None: 标题行
    sheet_index:int=0: 存储的excel页序号
    """
    workbook = openpyxl.Workbook()
    workbook.remove(workbook['Sheet'])
    new_sheet = workbook.create_sheet(sheet_name)
    first_data_row = 2 if title_row else 1
    if title_row:
        for title, col_num in zip(title_row, range(1, len(title_row)+1)):
            new_sheet.cell(row=1, column=col_num).value = title
    for row, row_num in zip(data_seqence, range(first_data_row, len(data_seqence)+first_data_row)):
        for cell_content, col_num in zip(row, range(1, len(row)+1)):
            # new_sheet.cell(row=row_num, column=col_num).value = ILLEGAL_CHARACTERS_RE.sub(r'',cell_content)
            new_sheet.cell(row=row_num, column=col_num).value = cell_content
    workbook.save(excel_file)
